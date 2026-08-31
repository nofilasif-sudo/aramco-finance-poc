"""Reading a table out of a worksheet a human authored.

Nothing in this module knows about charts of accounts, affiliates or bronze.
It solves one problem: a spreadsheet built for people to read has a title
block, a header somewhere below it, a body, and a footer — and none of those
boundaries are where you assume. Keeping that separate from the CoA rules is
what lets the same code read the trial balances, or next quarter's pack,
without being edited.
"""

from __future__ import annotations

import datetime as dt
import re

from openpyxl import load_workbook


class IngestError(Exception):
    """A source file did not look the way the contract says it should."""


# ---------------------------------------------------------------------------
# Cell primitives
# ---------------------------------------------------------------------------
def to_raw_str(value) -> str:
    """Render a cell as text, changing as little as possible.

    Only storage-format normalisation — undoing how xlsx chose to encode the
    value — never interpretation:

        None  -> ""
        1.0   -> "1"     xlsx has no integer type; the ".0" is the container's
                         artifact, not something anyone typed.
        1.5   -> "1.5"   float repr round-trips exactly, so this is lossless.
        date  -> "2024-03-31"   ISO; "3/31/24" is locale-ambiguous.

    Leading spaces are KEPT: the Group chart uses indentation to mark level-2
    sub-accounts, so that whitespace is data.
    """
    if value is None:
        return ""
    if isinstance(value, bool):           # before int: bool subclasses int
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return str(value)
        return str(int(value)) if value.is_integer() else repr(value)
    if isinstance(value, dt.datetime):
        return (value.date().isoformat() if value.time() == dt.time(0)
                else value.isoformat(sep=" "))
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def norm_header(value) -> str:
    """Normalise a header cell FOR MATCHING ONLY — never for output.

    Excel headers routinely carry a trailing space or an alt-enter line break
    that is invisible on screen. Matching the raw string makes the pipeline
    fail for a reason nobody can see.
    """
    return re.sub(r"\s+", " ", str(value if value is not None else "")).strip().lower()


def is_blank_row(row) -> bool:
    return all(c is None or str(c).strip() == "" for c in row)


def resolve_columns(headers: list[str], aliases: dict[str, list[str]],
                    required: set[str], where: str) -> dict[str, int]:
    """Map output column -> 0-based index in THIS sheet, via an alias table.

    Aliases rather than one exact string because the same concept is spelled
    differently across tabs — the CoA tabs say "Category (FS caption group)"
    where the TB tabs say "Category". Returns only the columns present; raises
    if a required one is missing, because a silently absent column lands an
    entire column of empty strings that looks like real "no value" data.
    """
    pos: dict[str, int] = {}
    for out_col, spellings in aliases.items():
        for spelling in spellings:
            if spelling in headers:
                pos[out_col] = headers.index(spelling)
                break
    missing = required - set(pos)
    if missing:
        raise IngestError(
            f"{where}: required column(s) {sorted(missing)} not found. "
            f"Headers seen: {[h for h in headers if h]}")
    return pos


# ---------------------------------------------------------------------------
# The table
# ---------------------------------------------------------------------------
class Table:
    """One parsed worksheet region: header, body, footer.

    All three boundaries are DISCOVERED:

      header : first row whose first cell is the sentinel. Every tab in this
               pack puts it on row 6, but a title row added above would shift
               the whole table and we would map names into the wrong column.
      body   : ends at the first fully blank row. (Verified against the pack:
               no blank rows occur inside a table, so this cannot truncate.)
      footer : every non-blank row after that. Not data, but not noise — it
               carries the "Total accounts: N" control total.
    """

    def __init__(self, title: str, rows: list[tuple], sentinel: str):
        self.title = title
        self.header_idx = self._find_header(rows, sentinel)
        # The title block above the header. Kept rather than discarded because
        # it is sometimes the ONLY place a fact appears — the CoA mapping tabs
        # state the affiliate's 4-digit code there and nowhere else.
        self.title_block = list(rows[:self.header_idx])
        self.raw_headers = list(rows[self.header_idx])
        self.headers = [norm_header(c) for c in self.raw_headers]

        self.data: list[tuple] = []
        self.footer: list[tuple] = []
        in_body = True
        for row in rows[self.header_idx + 1:]:
            if in_body:
                if is_blank_row(row):
                    in_body = False
                else:
                    self.data.append(row)
            elif not is_blank_row(row):
                self.footer.append(row)

    @staticmethod
    def _find_header(rows: list[tuple], sentinel: str) -> int:
        for i, row in enumerate(rows):
            if row and norm_header(row[0]) == sentinel:
                return i
        raise IngestError(f"no header row (first cell '{sentinel}') found")

    def footer_text(self) -> str:
        return " ".join(to_raw_str(c) for row in self.footer
                        for c in row if c is not None).strip()

    def title_text(self) -> str:
        """The title block flattened to one string, for pattern matching.

        Symmetric with footer_text(): the same "the fact is somewhere in this
        region, not in a known cell" problem, at the other end of the table.
        """
        return " ".join(to_raw_str(c) for row in self.title_block
                        for c in row if c is not None).strip()

    @staticmethod
    def classify(row: tuple) -> str:
        """account / section / subtotal / other.

        The sheets encode row kind positionally, not with a flag column:
            account  : code in col A AND name in col B
            section  : text in col A only   ("NON-CURRENT ASSETS")
            subtotal : text in col B only   ("  Subtotal - Revenue")
        Defined once here so the extractor and the control total can never
        disagree about what an account is.
        """
        a = to_raw_str(row[0] if len(row) > 0 else None).strip()
        b = to_raw_str(row[1] if len(row) > 1 else None).strip()
        if a and b:
            return "account"
        if a:
            return "section"
        if b:
            return "subtotal"
        return "other"

    def counts(self) -> dict[str, int]:
        out = {"account": 0, "section": 0, "subtotal": 0, "other": 0}
        for row in self.data:
            out[self.classify(row)] += 1
        return out


def read_tables(path, sentinel: str):
    """Yield (sheet_title, Table) for every tab that contains a header row.

    Tab selection is STRUCTURAL, not a name blacklist: 'Read me' and
    'Anchors & sources' are skipped because they have no header, so a new
    prose tab needs no code change.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            try:
                yield ws.title, Table(ws.title, rows, sentinel)
            except IngestError:
                continue
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Wide-table melting (trial-balance-shaped sheets: id columns, then one
# column per period). Shared by tb.py and group_tb.py so the two don't
# duplicate the same "everything right of the last id column is a period"
# rule twice.
# ---------------------------------------------------------------------------
def find_period_columns(raw_headers: list, last_id_idx: int) -> list[tuple[int, str]]:
    """Every non-empty header right of the last identifier column is a period.

    Not a hardcoded quarter list: next quarter adds a column, this sees it
    without a code change. Returns (column index, trimmed label) pairs.
    """
    periods = []
    for i, raw in enumerate(raw_headers):
        if i <= last_id_idx or norm_header(raw) == "":
            continue
        periods.append((i, to_raw_str(raw).strip()))
    return periods


# ---------------------------------------------------------------------------
# Generic footer / nil-proof checks for signed ledgers (trial balances).
# Only knows "sum the numeric cells in this column across rows classified as
# 'account'" — nothing here is TB-specific, so tb.py and group_tb.py share
# one implementation instead of two copies that could quietly disagree.
# ---------------------------------------------------------------------------
def sum_account_rows(tab: Table, col_idx: int) -> tuple[float, float]:
    """Return (signed sum, sum of magnitudes) of column `col_idx` over
    rows classified as 'account'. Magnitude is returned alongside the sum
    so a caller can size a relative tolerance around it."""
    total = magnitude = 0.0
    for row in tab.data:
        if Table.classify(row) != "account":
            continue
        v = row[col_idx] if col_idx < len(row) else None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            total += float(v)
            magnitude += abs(float(v))
    return total, magnitude


def check_row_by_label(tab: Table, contains: str):
    """Find the footer row whose 2nd cell contains `contains` (case-insensitive).

    Used to read a sheet's own proving row (e.g. "TRIAL BALANCE CHECK") as an
    independent second opinion alongside a value computed from the data.
    """
    contains = contains.lower()
    for row in tab.footer:
        b = to_raw_str(row[1] if len(row) > 1 else None)
        if contains in b.lower():
            return row
    return None
