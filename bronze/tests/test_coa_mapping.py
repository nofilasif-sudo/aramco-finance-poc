"""Tests for the CoA mapping ingestion.

Plain unittest, no pytest needed:  python -m unittest discover -s tests

These assert BEHAVIOUR, not shape. `assert not df.empty` passes when every
value is wrong; each test below fails if a specific decision regresses.

The load-bearing one is test_low_confidence_rows_still_land: the flagged and
unmapped rows are the deliverable, so an extractor that got "strict" and
refused them would break the demo while every other test still passed.
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from bronze_ingest.coa_mapping import (COLUMNS, affiliate_code, extract,
                                       write_csv)
from bronze_ingest.excel import IngestError

CFG = {
    "header_sentinel": "affiliate a/c",
    "tab_marker": "Fixture",
    "affiliate_pattern": r"Affiliate\s+(\d{4})",
    "csv_encoding": "utf-8",
}

# (account, name, node, node name, confidence, status, rationale)
AUTO = ("1100", "Land", "G11000", "PP&E (net)", 0.99, "Auto-mapped", None)
REVIEW = ("5020", "Catalysts & chemicals", "G52000", "Producing & mfg", 0.72,
          "Analyst review", "By-function vs by-nature ambiguity.")
UNMAPPED = ("8100", "Discontinued operations", "G57000", "Impairment", 0.35,
            "Unmapped - analyst intervention", "Group has no such caption.")


def fixture(rows=(AUTO, REVIEW, UNMAPPED), footer: str | None = None,
            title: str = "Affiliate 2010  ->  Aramco Group nodes",
            tab: str = "Mapping - Fixture", name: str = "map") -> Path:
    """A workbook shaped like the real pack, but deliberately different.

    The header sits on row 3 rather than row 5 and the workbook carries a
    second, non-matching mapping tab — so the test fails if header discovery
    regresses to a fixed row, or if tab selection stops filtering.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = tab
    ws.append(["CoA Mapping - Fixture -> Group"])
    ws.append([title])
    ws.append(["Affiliate a/c", "Affiliate account name", "Group node",
               "Group node name", "Confidence", "Status",
               "Rationale (flagged rows)"])
    for row in rows:
        ws.append(list(row))
    ws.append([])                                    # blank ends the body
    if footer is None:
        counts = {"Auto-mapped": 0, "Analyst review": 0,
                  "Unmapped - analyst intervention": 0}
        for row in rows:
            counts[row[5]] = counts.get(row[5], 0) + 1
        footer = (f"Auto-mapped: {counts['Auto-mapped']}    "
                  f"Analyst review: {counts['Analyst review']}    "
                  f"Unmapped: {counts['Unmapped - analyst intervention']}    "
                  f"Total: {len(rows)}")
    ws.append([None, footer])
    ws.row_dimensions[len(rows) + 4].height = 15  # force the blank into the XML

    # A second mapping tab that tab_marker must NOT select.
    other = wb.create_sheet("Mapping - Other")
    other.append(["CoA Mapping - Other -> Group"])
    other.append(["Affiliate 9999  ->  Aramco Group nodes"])
    other.append(["Affiliate a/c", "Affiliate account name", "Group node",
                  "Group node name", "Confidence", "Status",
                  "Rationale (flagged rows)"])
    other.append(list(AUTO))
    other.append([])
    other.append([None, "Auto-mapped: 1    Analyst review: 0    "
                        "Unmapped: 0    Total: 1"])
    other.row_dimensions[5].height = 15

    path = Path(tempfile.gettempdir()) / f"coa_mapping_{name}.xlsx"
    wb.save(path)
    return path


class TestAffiliateCode(unittest.TestCase):
    def test_code_comes_from_the_title_block_not_the_tab_name(self):
        # The tab name carries the trade name; only the title block carries
        # the 4-digit code the rest of the warehouse joins on.
        self.assertEqual(
            affiliate_code("Mapping - SABIC",
                           "Affiliate 2010  ->  Aramco Group nodes",
                           r"Affiliate\s+(\d{4})"), "2010")

    def test_unattributable_tab_raises(self):
        # Landing 66 mappings with a blank affiliate_code would make them
        # unjoinable — 2010 and 2380 share 4-digit account codes.
        with self.assertRaises(IngestError):
            affiliate_code("Mapping - Mystery", "no code here",
                           r"Affiliate\s+(\d{4})")


class TestExtract(unittest.TestCase):
    def test_header_found_not_assumed_and_aliases_resolve(self):
        rows, _ = extract(fixture(), CFG, [])
        self.assertEqual([r["affiliate_account"] for r in rows],
                         ["1100", "5020", "8100"])
        # 'Rationale (flagged rows)' must resolve to `rationale`
        self.assertEqual(rows[1]["rationale"],
                         "By-function vs by-nature ambiguity.")

    def test_affiliate_code_added_to_every_row(self):
        rows, _ = extract(fixture(), CFG, [])
        self.assertTrue(all(r["affiliate_code"] == "2010" for r in rows))

    def test_source_file_populated_from_workbook_name(self):
        path = fixture()
        rows, _ = extract(path, CFG, [])
        self.assertTrue(all(r["source_file"] == path.name for r in rows))

    def test_only_the_marked_tab_is_read(self):
        rows, meta = extract(fixture(), CFG, [])
        self.assertEqual(list(meta), ["Mapping - Fixture"])
        self.assertNotIn("9999", {r["affiliate_code"] for r in rows})

    def test_a_marker_matching_no_tab_raises(self):
        with self.assertRaises(IngestError):
            extract(fixture(), {**CFG, "tab_marker": "Nonexistent"}, [])

    def test_a_marker_matching_two_tabs_raises(self):
        # 'Mapping' matches both tabs; landing both would put the same rows in
        # both tables.
        with self.assertRaises(IngestError):
            extract(fixture(), {**CFG, "tab_marker": "Mapping"}, [])

    def test_confidence_kept_as_written_not_rounded(self):
        rows, _ = extract(fixture(), CFG, [])
        self.assertEqual(rows[0]["confidence"], "0.99")
        self.assertEqual(rows[2]["confidence"], "0.35")

    def test_empty_rationale_lands_as_empty_string(self):
        # Not "\\N": rationale is NOT NULL in the DDL, unlike fs_*.note_ref.
        rows, _ = extract(fixture(), CFG, [])
        self.assertEqual(rows[0]["rationale"], "")


class TestFlaggedRowsAreTheDeliverable(unittest.TestCase):
    def test_low_confidence_rows_still_land(self):
        """The point of the whole table. An extractor that refused a 0.35
        score would be unable to ingest the row the demo is built around."""
        rows, _ = extract(fixture(), CFG, [])
        unmapped = [r for r in rows if r["status"].startswith("Unmapped")]
        self.assertEqual(len(unmapped), 1)
        self.assertEqual(unmapped[0]["affiliate_account"], "8100")
        self.assertEqual(unmapped[0]["confidence"], "0.35")

    def test_status_counts_reported_per_tab(self):
        _, meta = extract(fixture(), CFG, [])
        self.assertEqual(meta["Mapping - Fixture"]["status_counts"],
                         {"Auto-mapped": 1, "Analyst review": 1,
                          "Unmapped - analyst intervention": 1})


class TestChecks(unittest.TestCase):
    def test_control_total_mismatch_refuses_to_land(self):
        # A check that has never been seen to fail is not a check.
        with self.assertRaises(IngestError):
            extract(fixture(footer="Auto-mapped: 1    Analyst review: 1    "
                                   "Unmapped: 1    Total: 99", name="badtotal"),
                    CFG, [])

    def test_per_status_mismatch_refuses_to_land(self):
        # Total is right but the bands are not — this is what catches `status`
        # being read from the wrong column index.
        with self.assertRaises(IngestError):
            extract(fixture(footer="Auto-mapped: 3    Analyst review: 0    "
                                   "Unmapped: 0    Total: 3", name="badband"),
                    CFG, [])

    def test_status_contradicting_its_confidence_refuses_to_land(self):
        bad = ("1100", "Land", "G11000", "PP&E", 0.20, "Auto-mapped", None)
        with self.assertRaises(IngestError):
            extract(fixture(rows=(bad,), name="badstatus"), CFG, [])

    def test_unknown_status_refuses_to_land(self):
        bad = ("1100", "Land", "G11000", "PP&E", 0.99, "Probably fine?", None)
        with self.assertRaises(IngestError):
            extract(fixture(rows=(bad,), name="unknownstatus"), CFG, [])

    def test_unparseable_confidence_refuses_to_land(self):
        bad = ("1100", "Land", "G11000", "PP&E", "high", "Auto-mapped", None)
        with self.assertRaises(IngestError):
            extract(fixture(rows=(bad,), name="badconf"), CFG, [])

    def test_blank_group_node_refuses_to_land(self):
        bad = ("1100", "Land", None, "PP&E", 0.99, "Auto-mapped", None)
        with self.assertRaises(IngestError):
            extract(fixture(rows=(bad,), name="blanknode"), CFG, [])

    def test_duplicate_affiliate_account_refuses_to_land(self):
        with self.assertRaises(IngestError):
            extract(fixture(rows=(AUTO, AUTO), name="dupe"), CFG, [])

    def test_expected_rows_mismatch_refuses_to_land(self):
        with self.assertRaises(IngestError):
            extract(fixture(name="expected"), {**CFG, "expected_rows": 99}, [])

    def test_missing_required_column_refuses_to_land(self):
        # A silently absent column lands a whole column of empty strings that
        # looks like real "no value" data.
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapping - Fixture"
        ws.append(["CoA Mapping - Fixture -> Group"])
        ws.append(["Affiliate 2010  ->  Aramco Group nodes"])
        ws.append(["Affiliate a/c", "Affiliate account name", "Group node",
                   "Group node name", "Confidence", "Status"])   # no rationale
        ws.append(list(AUTO[:6]))
        path = Path(tempfile.gettempdir()) / "coa_mapping_nocol.xlsx"
        wb.save(path)
        with self.assertRaises(IngestError):
            extract(path, CFG, [])


class TestSink(unittest.TestCase):
    def test_header_matches_the_schema_exactly(self):
        rows, _ = extract(fixture(), CFG, [])
        out = Path(tempfile.gettempdir()) / "coa_mapping_test.csv"
        write_csv(rows, out, "utf-8", [])
        first = out.read_text(encoding="utf-8").splitlines()[0]
        self.assertEqual(first, ",".join(COLUMNS))

    def test_line_endings_are_lf_on_every_platform(self):
        rows, _ = extract(fixture(), CFG, [])
        out = Path(tempfile.gettempdir()) / "coa_mapping_test.csv"
        write_csv(rows, out, "utf-8", [])
        self.assertNotIn(b"\r\n", out.read_bytes())


if __name__ == "__main__":
    unittest.main()
